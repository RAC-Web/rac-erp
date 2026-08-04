from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from apps.attendance.models import Attendance
from apps.leave.models import LeaveRequest
from apps.students.models import StudentProfile
from apps.payroll.models import PayrollRecord
from apps.conveyance.models import Conveyance
import io

@login_required
def reports_dashboard(request):
    if getattr(request.user, 'role', '') not in ('Principal', 'Manager'):
        return HttpResponse("Unauthorized", status=403)
    return render(request, 'reports/dashboard.html')

@login_required
def attendance_report(request):
    if getattr(request.user, 'role', '') not in ('Principal', 'Manager'):
        return HttpResponse("Unauthorized", status=403)
    
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    records = Attendance.objects.all().select_related('student').order_by('-date')
    
    if month and year:
        records = records.filter(date__month=int(month), date__year=int(year))
    
    return render(request, 'reports/attendance_report.html', {
        'records': records,
        'month': month,
        'year': year,
    })

@login_required
def attendance_report_excel(request):
    if getattr(request.user, 'role', '') not in ('Principal', 'Manager'):
        return HttpResponse("Unauthorized", status=403)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)
    
    month = request.GET.get('month')
    year = request.GET.get('year')
    
    records = Attendance.objects.all().select_related('student').order_by('date')
    if month and year:
        records = records.filter(date__month=int(month), date__year=int(year))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Rahman Anis & Co. — Attendance Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Student Name', 'Date', 'Time IN', 'Time OUT', 'Working Hours', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data
    for row_idx, record in enumerate(records, 4):
        ws.cell(row=row_idx, column=1, value=record.student.full_name).border = thin_border
        ws.cell(row=row_idx, column=2, value=record.date.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row_idx, column=3, value=record.time_in.strftime('%I:%M %p') if record.time_in else '-').border = thin_border
        ws.cell(row=row_idx, column=4, value=record.time_out.strftime('%I:%M %p') if record.time_out else '-').border = thin_border
        ws.cell(row=row_idx, column=5, value=str(record.working_hours) if record.working_hours else '-').border = thin_border
        ws.cell(row=row_idx, column=6, value=record.status).border = thin_border
    
    # Auto-size columns
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.xlsx"'
    return response

@login_required
def payroll_report(request):
    if getattr(request.user, 'role', '') not in ('Principal', 'Manager'):
        return HttpResponse("Unauthorized", status=403)
    
    records = PayrollRecord.objects.all().select_related('student').order_by('-month')
    return render(request, 'reports/payroll_report.html', {'records': records})

@login_required
def payroll_report_excel(request):
    if getattr(request.user, 'role', '') not in ('Principal', 'Manager'):
        return HttpResponse("Unauthorized", status=403)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)
    
    records = PayrollRecord.objects.all().select_related('student').order_by('-month')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:H1')
    ws['A1'] = "Rahman Anis & Co. — Payroll Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Student', 'Month', 'Base Stipend', 'Conveyance', 'Bonus', 'Late Ded.', 'Leave Ded.', 'Net Pay']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row_idx, record in enumerate(records, 4):
        ws.cell(row=row_idx, column=1, value=record.student.full_name).border = thin_border
        ws.cell(row=row_idx, column=2, value=record.month.strftime('%b %Y')).border = thin_border
        ws.cell(row=row_idx, column=3, value=float(record.base_stipend)).border = thin_border
        ws.cell(row=row_idx, column=4, value=float(record.conveyance_allowance)).border = thin_border
        ws.cell(row=row_idx, column=5, value=float(record.bonus)).border = thin_border
        ws.cell(row=row_idx, column=6, value=float(record.late_deduction)).border = thin_border
        ws.cell(row=row_idx, column=7, value=float(record.leave_deduction)).border = thin_border
        ws.cell(row=row_idx, column=8, value=float(record.net_pay)).border = thin_border
    
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 16
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payroll_report.xlsx"'
    return response
